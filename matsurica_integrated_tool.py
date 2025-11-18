# -*- coding: utf-8 -*-
# 最終更新: 2025-11-17 14:25 (Codexによる追記)
"""
マツリカ統合ツール
Excelファイルからマツリカ取込用CSVを生成する統合ツール

処理フロー:
1. ExcelファイルをCSVに変換
2. 顧客リストとマッチング
3. マツリカ取込用CSVを生成
"""

import argparse
import sys
from pathlib import Path
import pandas as pd
import zipfile
import xml.etree.ElementTree as ET
import re
import unicodedata
from datetime import datetime

TOOL_VERSION = "matsurica_integrated_tool.py v2025.10.22-02"

CUSTOMER_NAME_ALIASES = [
    "取引先名(必須)",
    "取引先名",
    "顧客名",
    "会社名",
    "企業名",
]

CUSTOMER_ID_ALIASES = [
    "取引先ID(必須)",
    "取引先ID",
    "顧客ID",
    "会社ID",
    "顧客コード",
    "取引先コード",
]

CUSTOMER_KUBUN_ALIASES = [
    "顧客区分（管理番号:19103）",
    "顧客区分（管理番号：19103）",
    "顧客区分",
    "顧客ランク",
    "区分",
]

CUSTOMER_MA_SUPPORT_ALIASES = [
    "MA部支援担当（管理番号:19258）",
    "MA部支援担当（管理番号：19258）",
    "MA部支援担当",
    "支援担当者",
    "担当者",
]

# ========== 共通ユーティリティ関数 ==========

def to_sjis_safe(s):
    """
    cp932で安全にエンコードできる文字列に変換する
    """
    if s is None:
        return ""
    if not isinstance(s, str):
        s = str(s)
    try:
        s.encode("cp932", errors="strict")
        return s
    except Exception:
        return s.encode("cp932", errors="replace").decode("cp932", errors="ignore")

def clean_newlines(s: str) -> str:
    if not isinstance(s, str):
        return ""
    t = s.replace("_x000D_", "")
    t = t.replace("\r\n", "\n").replace("\r", "\n")
    t = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", "", t)
    return t

def find_col(df: pd.DataFrame, names_or_idx, default=None):
    """names_or_idx: list[str] or int or list[int]"""
    cols = df.columns.tolist()
    # by name (first match)
    if isinstance(names_or_idx, (list, tuple)):
        for name in names_or_idx:
            if isinstance(name, str) and name in cols:
                return name
            if isinstance(name, int) and 0 <= name < len(cols):
                return cols[name]
    elif isinstance(names_or_idx, str):
        return names_or_idx if names_or_idx in cols else default
    elif isinstance(names_or_idx, int):
        return cols[names_or_idx] if 0 <= names_or_idx < len(cols) else default
    return default

def normalize_column_label(label: str) -> str:
    """
    列名を比較しやすいように正規化する
    - 全角/半角を揃える
    - 括弧内の注釈（例: (必須)）を除去
    - スペースや区切り記号を除去
    """
    if label is None:
        return ""
    text = unicodedata.normalize("NFKC", str(label))
    # Remove annotations such as (必須) / （必須）
    text = re.sub(r"[（(][^）)]*[）)]", "", text)
    # Remove obvious keywords that are only annotations
    text = text.replace("必須", "")
    # Remove separators/whitespace
    text = re.sub(r"[\s　:_：・･／/、，,.-]", "", text)
    return text.strip()

def build_column_lookup(columns):
    """
    normalize_column_label をキーにした辞書を作成する
    """
    lookup = {}
    for col in columns:
        norm = normalize_column_label(col)
        if norm and norm not in lookup:
            lookup[norm] = col
    return lookup

def resolve_column(lookup_dict, aliases):
    """
    alias候補（文字列またはリスト）から最初に一致した列名を返す
    """
    if not isinstance(aliases, (list, tuple)):
        aliases = [aliases]
    for cand in aliases:
        norm = normalize_column_label(cand)
        if norm and norm in lookup_dict:
            return lookup_dict[norm]
    return None

# ========== Excel→CSV変換機能 ==========

def extract_shared_strings(excel_path):
    """
    共有文字列テーブルを抽出する
    """
    shared_strings = {}
    try:
        with zipfile.ZipFile(excel_path, 'r') as z:
            if 'xl/sharedStrings.xml' in z.namelist():
                with z.open('xl/sharedStrings.xml') as f:
                    content = f.read().decode('utf-8', errors='ignore')
                    # 簡易的なXMLパース（正規表現使用）
                    import re
                    string_matches = re.findall(r'<si>.*?<t[^>]*>(.*?)</t>.*?</si>', content, re.DOTALL)
                    for idx, string_val in enumerate(string_matches):
                        shared_strings[idx] = string_val
                    print(f"共有文字列を抽出: {len(shared_strings)}個")
            else:
                print("共有文字列テーブルが見つかりません")
    except Exception as e:
        print(f"共有文字列抽出エラー: {e}")
    
    return shared_strings

def get_sheet_mapping(excel_path):
    """
    workbook.xmlからシート名とファイルの対応関係を取得する
    """
    sheet_mapping = {}
    try:
        with zipfile.ZipFile(excel_path, 'r') as z:
            if 'xl/workbook.xml' in z.namelist():
                with z.open('xl/workbook.xml') as f:
                    content = f.read().decode('utf-8', errors='ignore')
                    # シート情報を正規表現で抽出
                    import re
                    sheet_matches = re.findall(r'<sheet name="([^"]+)" sheetId="(\d+)" r:id="([^"]+)"/>', content)
                    for name, sheet_id, r_id in sheet_matches:
                        sheet_mapping[name] = f"xl/worksheets/sheet{sheet_id}.xml"
                        print(f"シート発見: {name} -> sheet{sheet_id}.xml")
            else:
                print("workbook.xmlが見つかりません")
    except Exception as e:
        print(f"シートマッピング取得エラー: {e}")
    
    return sheet_mapping

def extract_sheet_data_from_zip(excel_path, target_sheet_name="明細データ"):
    """
    ZIPとしてExcelファイルを開き、指定されたシートのデータを直接抽出する
    """
    try:
        # まず共有文字列を取得
        shared_strings = extract_shared_strings(excel_path)
        
        with zipfile.ZipFile(excel_path, 'r') as z:
            # シートマッピングを取得
            sheet_mapping = get_sheet_mapping(excel_path)
            
            # 対象シートを探す
            target_sheet_file = None
            if target_sheet_name in sheet_mapping:
                target_sheet_file = sheet_mapping[target_sheet_name]
            else:
                # 明細データシートが見つからない場合、最初のシートを使用
                sheet_files = [name for name in z.namelist() if name.startswith('xl/worksheets/sheet') and name.endswith('.xml')]
                if sheet_files:
                    target_sheet_file = sheet_files[0]
                    print(f"明細データシートが見つからないため、最初のシートを使用: {target_sheet_file}")
            
            if not target_sheet_file:
                raise ValueError("シートファイルが見つかりません")
            
            print(f"シートファイルを抽出: {target_sheet_file} ({target_sheet_name})")
            
            # シートデータを読み込み
            with z.open(target_sheet_file) as f:
                sheet_content = f.read().decode('utf-8', errors='ignore')
            
            return parse_sheet_xml(sheet_content, shared_strings)
            
    except Exception as e:
        raise Exception(f"ZIPからのデータ抽出失敗: {e}")

def parse_sheet_xml(xml_content, shared_strings=None):
    """
    シートのXMLからデータを解析する
    """
    if shared_strings is None:
        shared_strings = {}
    
    try:
        # XML名前空間
        ns = {'ss': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        
        # XMLをパース
        root = ET.fromstring(xml_content)
        
        rows = []
        for row_elem in root.findall('.//ss:row', ns):
            row_data = []
            for cell_elem in row_elem.findall('.//ss:c', ns):
                # セルの値とデータ型を取得
                value_elem = cell_elem.find('.//ss:v', ns)
                if value_elem is not None:
                    cell_value = value_elem.text
                    # データ型に応じた処理
                    cell_type = cell_elem.get('t')
                    if cell_type == 's':  # 共有文字列
                        # 共有文字列テーブルからの参照を実際の文字列に変換
                        try:
                            str_index = int(cell_value)
                            if str_index in shared_strings:
                                cell_value = shared_strings[str_index]
                            else:
                                cell_value = f"STRING_{cell_value}"
                        except ValueError:
                            cell_value = f"STRING_{cell_value}"
                    row_data.append(cell_value)
                else:
                    row_data.append("")
            rows.append(row_data)
        
        return rows
        
    except ET.ParseError:
        # XMLパースエラーの場合、正規表現で簡易的にデータ抽出
        return extract_data_with_regex(xml_content, shared_strings)

def extract_data_with_regex(xml_content, shared_strings=None):
    """
    正規表現で強制的にデータを抽出する（XMLパース失敗時のフォールバック）
    """
    if shared_strings is None:
        shared_strings = {}
    
    print("XMLパース失敗、正規表現でデータ抽出を試みます...")
    
    rows = []
    # 行を検出
    row_pattern = r'<row[^>]*>(.*?)</row>'
    cell_pattern = r'<c[^>]*>(.*?)</c>'
    value_pattern = r'<v[^>]*>(.*?)</v>'
    type_pattern = r't="([^"]*)"'
    
    row_matches = re.findall(row_pattern, xml_content, re.DOTALL)
    for row_match in row_matches:
        row_data = []
        cell_matches = re.findall(cell_pattern, row_match, re.DOTALL)
        for cell_match in cell_matches:
            value_match = re.search(value_pattern, cell_match, re.DOTALL)
            type_match = re.search(type_pattern, cell_match)
            
            cell_value = ""
            if value_match:
                cell_value = value_match.group(1)
                # データ型に応じた処理
                if type_match and type_match.group(1) == 's':
                    # 共有文字列の参照を実際の文字列に変換
                    try:
                        str_index = int(cell_value)
                        if str_index in shared_strings:
                            cell_value = shared_strings[str_index]
                        else:
                            cell_value = f"STRING_{cell_value}"
                    except ValueError:
                        cell_value = f"STRING_{cell_value}"
            
            row_data.append(cell_value)
        rows.append(row_data)
    
    return rows

def force_excel_to_csv(input_excel_path, output_csv_path=None, target_sheet_name="明細データ"):
    """
    破損したExcelファイルを強制的にCSVに変換
    target_sheet_name: 処理対象のシート名（デフォルトは"明細データ"）
    """
    input_path = Path(input_excel_path)
    if not input_path.exists():
        raise FileNotFoundError(f"入力ファイルが見つかりません: {input_excel_path}")
    
    if output_csv_path is None:
        output_csv_path = input_path.with_suffix('.csv')
    
    print(f"強制変換開始: {input_excel_path}")
    print(f"対象シート: {target_sheet_name}")
    
    try:
        # まず通常の方法を試す（明細データシートを指定）
        try:
            df = pd.read_excel(input_excel_path, sheet_name=target_sheet_name)
            print("通常の方法で読み込み成功")
        except Exception as e:
            print(f"通常の方法失敗: {e}")
            # シートが見つからない場合は最初のシートを試す
            try:
                df = pd.read_excel(input_excel_path, sheet_name=0)
                print("最初のシートで読み込み成功")
            except Exception as e2:
                print(f"最初のシートも失敗: {e2}")
                # 強制変換を実行（明細データシートを指定）
                data = extract_sheet_data_from_zip(input_excel_path, target_sheet_name)
                print(f"抽出したデータ: {len(data)}行, サンプル: {data[:2] if len(data) > 1 else data}")
                
                # データフレーム作成（最初の行をヘッダーとして使用）
                if len(data) > 1:
                    df = pd.DataFrame(data[1:], columns=data[0])
                else:
                    df = pd.DataFrame(data)
                
                print(f"作成したDataFrame: {df.shape}")
                print(f"列の型: {type(df.columns)}")
                print(f"最初の数列: {df.columns[:5].tolist() if hasattr(df.columns, 'tolist') else list(df.columns)[:5]}")
        
        # 日付列（6列目）をExcelシリアル値から日付に変換
        if len(df.columns) > 5:
            date_col = df.columns[5]  # 6列目（0-indexedなので5） - 「活動日」列
            print(f"日付列を変換中: {date_col}")
            
            # 日付列が空の場合でも、列インデックスで特定して変換を試みる
            try:
                # 列名が空でも、列インデックスでアクセスできる
                date_series = df.iloc[:, 5]  # 6列目（0-indexedなので5） - 「活動日」列
                print(f"日付列のデータ型: {date_series.dtype}")
                print(f"変換前のサンプル: {date_series.head(5).values.tolist()}")
                
                # 日付列がExcelシリアル値かどうかをチェックして変換
                df.iloc[:, 5] = date_series.apply(convert_excel_serial_to_date)
                
                # 変換結果を確認
                sample_dates = df.iloc[:, 5].head(5).values.tolist()
                print(f"日付変換サンプル: {sample_dates}")
            except Exception as e:
                print(f"日付変換エラー: {e}")
        
        # 文字列データを安全に変換
        for col in df.columns:
            col_series = df[col]
            if hasattr(col_series, 'dtype') and col_series.dtype == 'object':
                df[col] = col_series.apply(to_sjis_safe)
        
        # CSV保存（cp932で強制保存、非対応文字は無視）
        df.to_csv(output_csv_path, index=False, encoding='cp932', errors='ignore')
        
        print(f"強制変換完了: {input_excel_path} -> {output_csv_path}")
        print(f"行数: {len(df)}, 列数: {len(df.columns)}")
        
        return output_csv_path
        
    except Exception as e:
        raise Exception(f"強制変換失敗: {e}")

# ========== 企業マッチング機能 ==========

# 正規化用正規表現
COMPANY_SUFFIX_RE = re.compile(
    r"(株式会社|（株）|\(株\)|㈱|有限会社|合同会社|合名会社|合資会社|"
    r"Co\.,?\s*Ltd\.?|Corporation|Company|Inc\.?)",
    flags=re.IGNORECASE,
)
PUNCT_RE = re.compile(r"[ \t\u3000‐\-–—・/／\.,，、\(\)\[\]{}<>『』「」" + r"”\"'’`･_＋\\+]+")
DIGIT_RE = re.compile(r"[0-9０-９]+")

def to_hiragana(s: str) -> str:
    return "".join([chr(ord(c)-0x60) if 'ァ' <= c <= 'ヴ' else c for c in s])

def normalize_text(s: str) -> str:
    """会社種類/空白/記号/数字を落とし、全半角統一・英字小文字・カナはひらがな化"""
    if not isinstance(s, str) or not s.strip():
        return ""
    t = unicodedata.normalize("NFKC", s)
    t = COMPANY_SUFFIX_RE.sub("", t)
    t = PUNCT_RE.sub("", t)
    t = DIGIT_RE.sub("", t)
    t = t.lower()
    t = to_hiragana(t)
    return t

def pick_col(df: pd.DataFrame, desired_name: str, fallback_idx: int):
    if desired_name in df.columns:
        return desired_name
    if 0 <= fallback_idx < len(df.columns):
        return df.columns[fallback_idx]
    return None

def remove_non_header_rows(df: pd.DataFrame) -> pd.DataFrame:
    """
    データフレームの先頭からヘッダーでない行を検出し削除する
    """
    if len(df) <= 1:
        return df
    
    original_length = len(df)
    rows_to_remove = 0
    found_header_row = -1
    
    # 最大5行までチェック（通常は1-2行のヘッダー想定）
    for check_row in range(min(5, len(df) - 1)):
        current_row = df.iloc[check_row]
        next_row = df.iloc[check_row + 1]
        
        # 現在の行がヘッダーらしいかチェック
        is_header_likely = False
        
        # 1. 現在の行の値のほとんどが文字列かチェック
        string_count = sum(1 for val in current_row if isinstance(val, str) and val and str(val).strip())
        if string_count / len(current_row) >= 0.6:  # 60%以上が文字列
            is_header_likely = True
        
        # 2. 典型的なヘッダーキーワードを含むかチェック
        header_keywords = ['no', '案件', '番号', '活動', '活動先', '案件名', '活動日', '活動者', '組織', '行動', '種別', '実施', '内容', 'id', 'name', 'date', 'time']
        keyword_match = any(
            any(keyword in str(val).lower() for keyword in header_keywords if isinstance(val, str) and val)
            for val in current_row
        )
        if keyword_match:
            is_header_likely = True
        
        # 3. 次の行とのデータ型が異なるかチェック
        type_mismatch = 0
        total_comparable = 0
        for i, (val1, val2) in enumerate(zip(current_row, next_row)):
            # Noneや空文字はスキップ
            if pd.isna(val1) or val1 == "" or pd.isna(val2) or val2 == "":
                continue
            total_comparable += 1
            # 型が異なる場合
            if type(val1) != type(val2):
                type_mismatch += 1
        
        if total_comparable > 0 and type_mismatch / total_comparable >= 0.4:  # 40%以上で型が異なる
            is_header_likely = True
        
        # 4. 現在の行がデータ行らしいかチェック（数値や日付が多い）
        data_like_count = sum(1 for val in current_row 
                             if isinstance(val, (int, float)) or 
                                (isinstance(val, str) and any(c.isdigit() for c in str(val))))
        if data_like_count / len(current_row) >= 0.5:  # 50%以上がデータらしい
            is_header_likely = False
        
        # ヘッダーらしくなければこの行を削除対象に追加
        if not is_header_likely:
            rows_to_remove += 1
            print(f"警告: 行 {check_row + 1} がヘッダーでないため削除対象とします")
        else:
            # ヘッダーらしい行が見つかったら終了
            found_header_row = check_row
            break
    
    # ヘッダー行が見つかった場合、その行をカラム名として設定
    if found_header_row >= 0:
        print(f"完了: 行 {found_header_row + 1} をヘッダーとして設定します")
        # ヘッダー行をカラム名として設定
        header_values = []
        for val in df.iloc[found_header_row]:
            if pd.isna(val) or val == "":
                header_values.append("")
            else:
                header_values.append(str(val))
        df.columns = header_values
        
        # ヘッダー行以降のデータを保持
        df = df.iloc[found_header_row + 1:].reset_index(drop=True)
    elif rows_to_remove > 0:
        # ヘッダー行が見つからず、削除対象行がある場合
        print(f"警告: 先頭 {rows_to_remove} 行を削除します")
        df = df.iloc[rows_to_remove:].reset_index(drop=True)
    
    if len(df) < original_length:
        print(f"完了: 有効なデータ行数: {len(df)}/{original_length}")
    
    return df

def read_activity_robust(activity_path: Path) -> tuple[pd.DataFrame, str]:
    """
    CSVまたはExcelファイルを読み込み、データを抽出する
    1) CSVファイルの場合: pd.read_csv で読み込み
    2) Excelファイルの場合: 通常の読み込みとフォールバック処理
    3) ヘッダーチェックを行い、不要な行を削除
    """
    if activity_path.suffix.lower() == '.csv':
        # CSVファイルの場合
        try:
            df = pd.read_csv(activity_path, encoding='cp932')
            # ヘッダーチェックと不要行削除
            df = remove_non_header_rows(df)
            return df, "CSVデータ"
        except Exception as e:
            raise RuntimeError(f"CSVファイルの読み込みに失敗しました: {e}")
    else:
        # Excelファイルの場合
        # まず通常経路
        try:
            xls = pd.ExcelFile(activity_path, engine="openpyxl")
            sheet = "明細データ" if "明細データ" in xls.sheet_names else xls.sheet_names[0]
            df = pd.read_excel(activity_path, sheet_name=sheet, engine="openpyxl")
            
            # ヘッダーチェックと不要行削除
            df = remove_non_header_rows(df)
            return df, sheet
        except Exception as e:
            err = str(e)

        # フォールバック：openpyxl で値のみ吸い出し
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename=activity_path, read_only=True, data_only=True)  # ←値のみ
            sheet = "明細データ" if "明細データ" in wb.sheetnames else wb.sheetnames[0]
            ws = wb[sheet]

            # 先頭行をヘッダとみなす（空なら A,B,C.. を仮ヘッダ）
            rows = ws.iter_rows(values_only=True)
            first = next(rows, None)
            if first is None:
                raise ValueError("シートが空です。")

            # 先頭行に None ばかりなら仮ヘッダ
            if all(v is None or str(v).strip() == "" for v in first):
                width = ws.max_column or 1
                cols = [f"col_{i+1}" for i in range(width)]
                data = []
                for r in rows:
                    data.append(list(r) + [""]*(width - len(r)))
                df = pd.DataFrame(data, columns=cols)
            else:
                cols = [str(c) if c is not None else "" for c in first]
                data = []
                for r in rows:
                    # 行長が短い場合に合わせる
                    row_vals = list(r)
                    if len(row_vals) < len(cols):
                        row_vals += [""] * (len(cols) - len(row_vals))
                    data.append(row_vals[:len(cols)])
                df = pd.DataFrame(data, columns=cols)

            # ヘッダーチェックと不要行削除
            df = remove_non_header_rows(df)
            return df, sheet
        except Exception as e2:
            raise RuntimeError(f"活動ファイルの読み込みに失敗しました。\n"
                               f"- 通常読込エラー: {err}\n"
                               f"- フォールバック読込エラー: {e2}")

def match_customers(customers_path: Path, activity_path: Path) -> pd.DataFrame:
    """
    顧客リストと活動ファイルをマッチングする
    """
    # 顧客リスト
    customers = pd.read_csv(customers_path, encoding="cp932")
    col_lookup = build_column_lookup(customers.columns)

    cust_name_col = resolve_column(col_lookup, CUSTOMER_NAME_ALIASES)
    if cust_name_col is None:
        available = ", ".join(map(str, customers.columns))
        raise RuntimeError(
            "顧客リストに必要な顧客名の列が見つかりません。\n"
            f"- 探索した候補: {', '.join(CUSTOMER_NAME_ALIASES)}\n"
            f"- 現在のヘッダー: {available}"
        )
    cust_id_col = resolve_column(col_lookup, CUSTOMER_ID_ALIASES)
    kubun_col   = resolve_column(col_lookup, CUSTOMER_KUBUN_ALIASES)

    # 活動ファイル（堅牢読込）
    activity, sheet = read_activity_robust(activity_path)

    # 列特定（C/M/G 相当）
    col_C = pick_col(activity, "活動先", 2)                 # C列
    col_M = activity.columns[12] if len(activity.columns) > 12 else None  # M列
    col_G = activity.columns[6]  if len(activity.columns) > 6  else None  # G列

    # 活動側 正規化文字列
    n = len(activity)
    norm_C = [normalize_text(v) if (col_C and isinstance(v, str)) else "" for v in (activity[col_C] if col_C else [""]*n)]
    norm_M = [normalize_text(v) if (col_M and isinstance(v, str)) else "" for v in (activity[col_M] if col_M else [""]*n)]
    norm_G = [normalize_text(v) if (col_G and isinstance(v, str)) else "" for v in (activity[col_G] if col_G else [""]*n)]

    # 出力列の準備
    matched_name = [None]*n
    matched_id   = [None]*n
    matched_kbn  = [None]*n
    remaining = set(range(n))

    # 顧客を1社ずつ → C→M→G の順に 未確定行へ 部分一致
    for _, crow in customers.iterrows():
        raw_name = str(crow.get(cust_name_col, "") or "")
        ckey = normalize_text(raw_name)
        if not ckey:
            continue

        to_fix = []
        for i in remaining:
            if ckey and ckey in norm_C[i]:
                to_fix.append(i)
        for i in list(remaining - set(to_fix)):
            if ckey and ckey in norm_M[i]:
                to_fix.append(i)
        for i in list(remaining - set(to_fix)):
            if ckey and ckey in norm_G[i]:
                to_fix.append(i)

        if to_fix:
            for i in to_fix:
                matched_name[i] = raw_name
                matched_id[i]   = crow.get(cust_id_col, None) if cust_id_col else None
                matched_kbn[i]  = crow.get(kubun_col, None)   if kubun_col  else None
            remaining -= set(to_fix)

        if not remaining:
            break

    # マッチングしなかった行を削除
    matched_indices = [i for i in range(n) if matched_name[i] is not None]
    unmatched_count = n - len(matched_indices)
    
    print(f"完了: マッチング結果: {len(matched_indices)}/{n} 行がマッチしました")
    if unmatched_count > 0:
        print(f"警告: {unmatched_count} 行がマッチしなかったため削除します")
    
    # マッチした行だけを保持
    out = activity.iloc[matched_indices].copy()
    out["マッチ顧客名"] = [matched_name[i] for i in matched_indices]
    out["取引先ID(必須)"] = [matched_id[i] for i in matched_indices]
    out["顧客区分（管理番号:19103）」"] = [matched_kbn[i] for i in matched_indices]  # 既存の表記ゆれに合わせておく

    return out

# ========== マツリカCSV生成機能 ==========

MGMT_HEAD_PAT = re.compile(r"^\s*■\s*(記入者|訪問日時|日時|提案機種|訪問者|販売店|訪問相手|顧客情報|活動ステージ)\s*[:：]?\s*$")

def extract_action_body_v6(text: str) -> str:
    """
    v6仕様: 活動内容から本文を抽出
    - 文頭に「：」を付けない
    - 「■活動内容」が含まれる場合はそれ以降のみ転記
    - 文頭の句読点やスペースを除去
    """
    if not isinstance(text, str):
        return "\"内容不明\""
    
    t = clean_newlines(text)
    
    # 「■活動内容」が含まれる場合、それ以降のみを抽出
    if "■活動内容" in t:
        parts = t.split("■活動内容", 1)
        if len(parts) > 1:
            body = parts[1].strip()
        else:
            body = t
    else:
        body = t
    
    # 管理見出しを削除
    lines = body.split("\n")
    clean_lines = []
    for line in lines:
        line_stripped = line.strip()
        if not MGMT_HEAD_PAT.match(line_stripped) and line_stripped:
            clean_lines.append(line_stripped)
    
    body = "\n".join(clean_lines)
    
    # 文頭の句読点やスペースを除去
    body = body.strip()
    if body:
        # 文頭が句読点や記号で始まる場合は除去
        while body and body[0] in "、。,.・:：;；!！?？\"'「」『』【】[]()（）":
            body = body[1:].strip()
    
    body = to_sjis_safe(body)
    
    if not body:
        body = "\"内容不明\""
    
    return body

# アクション種別判定用キーワード
MAIL_WORDS = ["送付", "返信", "メール", "送信", "添付", "cc", "エビデンス", "提出"]
PHONE_WORDS = ["架電", "折返", "通話", "連絡", "コール", "電話"]
OUT_TASK_WORDS = ["現調", "立会", "設置", "納品", "リモート設定", "現地", "フィールド", "調整"]
IN_TASK_WORDS = ["見積", "資料作成", "社内", "mtg", "整理", "手配", "稟議", "準備"]

def decide_action_type(method_val: str, k_type: str, free_text: str) -> str:
    val = "社外タスク"  # デフォルト値
    mv = str(method_val).strip() if method_val is not None else ""
    kv = str(k_type).strip() if k_type is not None else ""
    ft = str(free_text).lower() if free_text is not None else ""

    # 手段列が「対面」または行動種別が「対面」の場合は「面談」にマッピング（仕様書準拠）
    if mv == "対面" or kv == "対面":
        return "面談"

    # K相当（行動種別）から一次判定
    if "電話" in kv:
        val = "電話"
    elif "メール" in kv:
        val = "メール"
    elif "会議" in kv or "mtg" in kv.lower():
        val = "社内タスク"

    # フリーテキストから補強・上書き
    if any(w in ft for w in MAIL_WORDS):
        val = "メール"
    elif any(w in ft for w in PHONE_WORDS):
        val = "電話"
    elif any(w in ft for w in OUT_TASK_WORDS):
        val = "社外タスク"
    elif any(w in ft for w in IN_TASK_WORDS):
        val = "社内タスク"

    return val

# 日付・時間抽出用正規表現
DATE_PAT1 = (
    r"(?P<y>\d{4})\s*(?:[/-]|年)\s*(?P<m>\d{1,2})"
    r"\s*(?:[/-]|月)\s*(?P<d>\d{1,2})(?:日)?"
)
DATE_PAT2 = r"(?P<mj>\d{1,2})月(?P<dj>\d{1,2})日"
TIME_PAT1  = r"(?P<h1>\d{1,2})[:：時](?P<min1>\d{0,2})"
TIME_PAT2  = r"(?P<h2>\d{1,2})[:：時](?P<min2>\d{0,2})"
RANGE_SEP = r"[～〜~\-ー−—]"

DATE_TIME_RANGE = re.compile(
    rf"(?:{DATE_PAT1}|{DATE_PAT2}).*?(?:{TIME_PAT1})?\s*{RANGE_SEP}\s*(?:{TIME_PAT2})",
    flags=re.IGNORECASE
)
DATE_ONLY_WEST = re.compile(DATE_PAT1, flags=re.IGNORECASE)
DATE_ONLY_JP = re.compile(DATE_PAT2, flags=re.IGNORECASE)
TIME_SINGLE = re.compile(r"(?P<h>\d{1,2})[:：時](?P<min>\d{0,2})")

def is_valid_date(year: int, month: int, day: int) -> bool:
    """日付が有効かどうかをチェックする"""
    try:
        datetime(year, month, day)
        return True
    except ValueError:
        return False

def convert_excel_serial_to_date(serial_value):
    """Excelシリアル値を日付に変換"""
    try:
        # 文字列の場合は数値に変換を試みる
        if isinstance(serial_value, str) and serial_value.strip():
            try:
                serial_value = float(serial_value)
            except ValueError:
                return serial_value  # 変換できない場合は元の値を返す
        
        if isinstance(serial_value, (int, float)) and 0 < serial_value < 100000:
            # Excelシリアル値を日付に変換（1900年1月1日を基準）
            base_date = datetime(1900, 1, 1)
            # Excelのバグ補正（1900年を閏年と誤認しているため）
            if serial_value >= 60:
                serial_value -= 1
            target_date = base_date + pd.Timedelta(days=serial_value - 1)
            return target_date.strftime("%Y-%m-%d")
    except Exception:
        pass
    return serial_value

def parse_dt_range(text: str, fallback_date):
    start_date = end_date = None
    start_time = end_time = None

    if isinstance(text, str):
        t = clean_newlines(text)
        m = DATE_TIME_RANGE.search(t)
        if m:
            # 西暦年月日形式 (yyyy/mm/dd or yyyy-mm-dd)
            if m.groupdict().get("y"):
                y = int(m.group("y")); mo = int(m.group("m")); d = int(m.group("d"))
                if is_valid_date(y, mo, d):
                    start_date = datetime(y, mo, d).strftime("%Y-%m-%d")
                    end_date   = start_date
                else:
                    print(f"警告: 無効な日付を検出: {y}/{mo}/{d}")
            
            # 和暦月日形式 (mm月dd日)
            elif m.groupdict().get("mj"):
                y = fallback_date.year if hasattr(fallback_date, "year") else datetime.now().year
                mo = int(m.group("mj")); d = int(m.group("dj"))
                if is_valid_date(y, mo, d):
                    start_date = datetime(y, mo, d).strftime("%Y-%m-%d")
                    end_date   = start_date
                else:
                    print(f"警告: 無効な日付を検出: {y}年{mo}月{d}日")

            sh, sm = m.group("h1"), m.group("min1")
            eh, em = m.group("h2"), m.group("min2")
            if sh:
                sm = sm if sm else "00"
                start_time = f"{int(sh):02d}:{int(sm):02d}"
            if eh:
                em = em if em else "00"
                end_time = f"{int(eh):02d}:{int(em):02d}"
        else:
            # タイムレンジとしては取得できなかった場合でも、日付単体・時間単体を拾う
            date_match = DATE_ONLY_WEST.search(t)
            if date_match:
                y = int(date_match.group("y")); mo = int(date_match.group("m")); d = int(date_match.group("d"))
                if is_valid_date(y, mo, d):
                    start_date = datetime(y, mo, d).strftime("%Y-%m-%d")
                    end_date = start_date
                else:
                    print(f"警告: 無効な日付を検出: {y}/{mo}/{d}")
            else:
                jp_match = DATE_ONLY_JP.search(t)
                if jp_match:
                    y = fallback_date.year if hasattr(fallback_date, "year") else datetime.now().year
                    mo = int(jp_match.group("mj")); d = int(jp_match.group("dj"))
                    if is_valid_date(y, mo, d):
                        start_date = datetime(y, mo, d).strftime("%Y-%m-%d")
                        end_date = start_date
                    else:
                        print(f"警告: 無効な日付を検出: {y}年{mo}月{d}日")

            time_match = TIME_SINGLE.search(t)
            if time_match:
                mh = time_match.group("h")
                mm = time_match.group("min") or "00"
                start_time = f"{int(mh):02d}:{int(mm):02d}"

    # fallback
    try:
        fb = pd.to_datetime(fallback_date) if fallback_date is not None else None
    except Exception:
        fb = None

    if not start_date:
        start_date = (fb or datetime.now()).strftime("%Y-%m-%d")
    if not end_date:
        end_date = start_date
    if not start_time:
        start_time = "10:00"
    if not end_time:
        # 単一時刻のみ取得できた場合は終了時間を開始時間に合わせる
        end_time = start_time if start_time else "11:00"

    return start_date, start_time, end_date, end_time

def build_output(activity_df: pd.DataFrame, customers_df: pd.DataFrame, template_cols: list[str], customer_lookup: dict[str, str] | None = None) -> pd.DataFrame:
    cols = activity_df.columns.tolist()
    if customer_lookup is None:
        customer_lookup = build_column_lookup(customers_df.columns)
    cust_id_master_col = resolve_column(customer_lookup, CUSTOMER_ID_ALIASES)
    ma_support_col     = resolve_column(customer_lookup, CUSTOMER_MA_SUPPORT_ALIASES)

    # よくある列名の候補
    col_customer_name = find_col(activity_df, ["マッチ顧客名", "活動先", 2], default=None)
    col_customer_id   = find_col(activity_df, ["取引先ID(必須)"], default=None)
    col_kubun         = find_col(activity_df, ["顧客区分（管理番号:19103）", "顧客区分（管理番号:19103）」"], default=None)
    col_method        = find_col(activity_df, ["方法", "活動方法", "訪問方法", 4], default=None)
    col_k_type        = find_col(activity_df, ["活動種別", "カテゴリ", "行動種別", 10], default=None)
    col_body          = find_col(activity_df, ["活動内容", "実施内容", "内容", "備考", 11], default=None)
    col_date          = find_col(activity_df, ["日付", "活動日", "訪問日", 7], default=None)

    # 出力用の最低限フィールドを抽出
    out_rows = []
    customer_id_index: dict[str, int] = {}
    if cust_id_master_col:
        for idx, value in customers_df[cust_id_master_col].items():
            if pd.isna(value):
                continue
            key = str(value).strip()
            if key and key not in customer_id_index:
                customer_id_index[key] = idx

    for i, row in activity_df.iterrows():
        cust_name = str(row.get(col_customer_name, "")) if col_customer_name else ""
        cust_id_raw = row.get(col_customer_id, "")
        cust_id_key = ""
        if cust_id_raw is not None and not pd.isna(cust_id_raw):
            cust_id_key = str(cust_id_raw).strip()
        kubun     = row.get(col_kubun, "")
        method_v  = row.get(col_method, "")
        ktype_v   = row.get(col_k_type, "")
        body_raw  = row.get(col_body, "")
        date_fallback = row.get(col_date, "")

        body = extract_action_body_v6(body_raw)
        d1, t1, d2, t2 = parse_dt_range(str(body_raw) if body_raw is not None else "", date_fallback)

        action_type = decide_action_type(method_v, ktype_v, body_raw)

        # テンプレ列に沿って dict を作る。既知のキーがあれば詰め、未知は空。
        rec = {col: "" for col in template_cols}

        # 代表的マッピング（出力見本の列名に合わせて可能な限り格納）
        for key in template_cols:
            if key == "取引先ID(必須)":
                rec[key] = to_sjis_safe(cust_id_key)
            elif key == "アクション種別(必須)":
                rec[key] = to_sjis_safe(action_type)
            elif key == "開始日(必須)":
                rec[key] = d1
            elif key == "開始時間(必須)":
                rec[key] = t1
            elif key == "終了日(必須)":
                rec[key] = d2
            elif key == "終了時間(必須)":
                rec[key] = t2
            elif key in ("実施結果", "詳細", "活動内容", "本文", "メモ"):
                rec[key] = body
            elif key in ("主担当者(必須)", "他の担当者", "事前メモ", "ステータス(必須)", "アクションコンタクト(コンタクトID)"):
                # これらの列にはデフォルト値を設定
                if key == "ステータス(必須)":
                    rec[key] = "完了"  # デフォルトステータス
                elif key == "主担当者(必須)":
                    # 顧客リストからMA部支援担当を取得
                    ma_support = ""
                    if cust_id_key and cust_id_key in customer_id_index and ma_support_col:
                        match_idx = customer_id_index[cust_id_key]
                        ma_support = customers_df.at[match_idx, ma_support_col]
                        if pd.isna(ma_support):
                            ma_support = ""
                    rec[key] = to_sjis_safe(ma_support if ma_support else "担当者未設定")
                else:
                    rec[key] = ""  # 空文字を設定

        out_rows.append(rec)

    out_df = pd.DataFrame(out_rows, columns=template_cols)
    return out_df

def build_matsurica_csv(customers_path: Path, matched_activity_path: Path) -> pd.DataFrame:
    """
    マツリカ取込用CSVを生成する
    """
    # 読み込み
    customers = pd.read_csv(customers_path, encoding="cp932")
    customer_lookup = build_column_lookup(customers.columns)
    xl = pd.ExcelFile(matched_activity_path)
    sheet = "明細データ" if "明細データ" in xl.sheet_names else xl.sheet_names[0]
    activity = pd.read_excel(matched_activity_path, sheet_name=sheet)

    # 出力見本の列をハードコード（customer_action_import_format.csvの列構造）
    template_cols = [
        '取引先ID(必須)', 'アクション種別(必須)', '開始日(必須)', '開始時間(必須)', 
        '終了日(必須)', '終了時間(必須)', '主担当者(必須)', '他の担当者', 
        '事前メモ', '実施結果', 'ステータス(必須)', 'アクションコンタクト(コンタクトID)'
    ]

    out_df = build_output(activity, customers, template_cols, customer_lookup)
    return out_df

# ========== メイン処理 ==========

def main(*args, **kwargs):
    parser = argparse.ArgumentParser(description='マツリカ統合ツール: Excelファイルからマツリカ取込用CSVを生成')
    parser.add_argument('input_excel', help='入力Excelファイルのパス')
    parser.add_argument('--customers', default='顧客リスト.csv', help='顧客リストCSVファイルのパス')
    parser.add_argument('--sample', default='出力見本：customer_action_import_format.csv', help='出力見本CSVファイルのパス（互換性のため保持）')
    parser.add_argument('--output', help='出力CSVファイルのパス')
    
    args = parser.parse_args()
    
    try:
        print("=== マツリカ統合ツール処理開始 ===")
        print(f"入力ファイル: {args.input_excel}")
        print(f"顧客リスト: {args.customers}")
        print(f"出力ファイル: {args.output if args.output else 'customer_action_import_format.csv'}")
        print(f"コードバージョン: {TOOL_VERSION}")
        
        # 1. Excel→CSV変換（明細データシートを指定）
        print("1. ExcelファイルをCSVに変換中...")
        csv_path = force_excel_to_csv(args.input_excel, target_sheet_name="明細データ")
        print(f"CSV変換完了: {csv_path}")
        print(f"変換されたCSVファイル: {csv_path}")
        
        # 2. 企業マッチング
        print("2. 顧客リストとマッチング中...")
        customers_path = Path(args.customers)
        if not customers_path.exists():
            raise FileNotFoundError(f"顧客リストファイルが見つかりません: {customers_path}")
        
        print(f"顧客リストファイルを読み込み中: {customers_path}")
        matched_activity = match_customers(customers_path, Path(csv_path))
        matched_path = Path("matched_activity.xlsx")
        print(f"マッチング結果をExcelファイルに保存中: {matched_path}")
        with pd.ExcelWriter(matched_path, engine="xlsxwriter") as w:
            matched_activity.to_excel(w, index=False)
        print(f"完了: 企業マッチング完了: {matched_path}")
        print(f"マッチングされた行数: {len(matched_activity)}行")
        
        # 3. マツリカCSV生成
        print("3. マツリカ取込用CSV生成中...")
        output_df = build_matsurica_csv(customers_path, matched_path)
        print(f"マツリカCSV生成完了: {len(output_df)}行")
        
        # 出力ファイルパスの決定
        if args.output:
            output_path = Path(args.output)
        else:
            output_path = Path("customer_action_import_format.csv")
        
        print(f"出力ファイルに保存中: {output_path}")
        output_df.to_csv(output_path, index=False, encoding="cp932")
        print(f"完了: マツリカCSV生成完了: {output_path}")
        
        print("=== 処理が正常に完了しました ===")
        print(f"最終出力ファイル: {output_path.resolve()}")
        print(f"出力ファイルサイズ: {output_path.stat().st_size} bytes")
        
    except Exception as e:
        print(f"エラーが発生しました: {e}")
        import traceback
        print(f"エラー詳細:\n{traceback.format_exc()}")
        sys.exit(1)

if __name__ == "__main__":
    main()
